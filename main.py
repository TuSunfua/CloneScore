import threading
import openpyxl
from bs4 import BeautifulSoup
import requests
import os

num_thread = 6
destination = 'data'

subjects = ['Toán', 'Ngữ văn', 'Ngoại ngữ', 'Vật lý', 'Hóa học', 'Sinh học', 'Lịch sử', 'Địa lý', 'Giáo dục công dân']
province = {}

def load_province_data() -> None:
    global province
    if not os.path.exists('province.txt'):
        raise FileNotFoundError('province.txt not found')    
        
    with open('province.txt', 'r', encoding='utf-8') as file:
        province = eval(file.read())

def generate_url(sbd: str, year: int = 2024) -> str:
    return f'https://diemthi.vnexpress.net/index/detail/sbd/{sbd}/year/{year}'

def get_score(sbd: str):
    response = requests.get(generate_url(sbd))
    soup = BeautifulSoup(response.content, 'html.parser')

    tbody_tag = soup.select_one('.o-detail-thisinh .e-table tbody')
    
    if tbody_tag is None:
        return {}, False
    
    td_tags = tbody_tag.find_all('td')

    score: dict = {}
    for i in range(0, len(td_tags), 2):
        subject = td_tags[i].get_text(strip=True)
        score[subject] = float(td_tags[i + 1].get_text(strip=True))

    return score, True

def write_score_to_sheet_on_row(sheet, score: dict, sbd: str, row: int) -> None:
    sheet.cell(row=row, column=1, value=sbd)
    sheet.cell(row=row, column=2, value=province[sbd[:2]])
    for i in range(len(subjects)):
        if subjects[i] not in score:
            sheet.cell(row=row, column=i+3, value=None)
        else:
            sheet.cell(row=row, column=i+3, value=score[subjects[i]])

def create_works():
    # Vì dữ liệu của TP. Hà Nội và TP. Hồ Chí Minh lớn nên cho 2 thành phố này thành 2 jobs riêng biệt
    # và 61 tỉnh/thành còn lại sẽ do num_thead - 2 đảm nhiệm
    arr = list(province.keys())
    n = len(arr)

    if num_thread <= 0 or num_thread > n:
        raise ValueError('Invalid number of thread')
    
    size = (n - 2) // (num_thread - 2)
    remainder = (n - 2) % (num_thread - 2)
    
    result = [[arr[0]], [arr[1]]] # TP. Hà Nội và TP. Hồ Chí Minh

    start = 2
    for i in range(num_thread - 2):
        end = start + size
        if i < remainder:
            end += 1
        
        result.append(arr[start:end])
        start = end
    
    return result

def worker(works: list[str]) -> None:
    for work in works:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.cell(row=1, column=1, value='SBD')
        sheet.cell(row=1, column=2, value='Tỉnh/Thành phố')
        for i in range(len(subjects)):
            sheet.cell(row=1, column=i+3, value=subjects[i])

        last_sbd = int(work + 1) * 1000000 - 1
        curr_sbd = int(work) * 1000000 + 1
        curr_row = 2
        failed = 0

        while failed < 100 and curr_sbd <= last_sbd:
            sbd = str(curr_sbd).zfill(8)
            score, is_success = get_score(sbd)
            if not is_success:
                print(f'Failed to get score of {sbd}')
                failed += 1
            write_score_to_sheet_on_row(sheet, score, sbd, curr_row)
            curr_row += 1
            curr_sbd += 1

        workbook.save(f'{destination}/{work}.xlsx')
        print(f"Finished collecting data of {province[work]}")

def main():
    try:
        load_province_data()
        works = create_works()

        if not os.path.exists(destination):
            os.makedirs(destination)
        
        threads = []
        for i in range(num_thread):
            thread = threading.Thread(target=worker, args=(works[i],))
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

    except Exception as e:
        print('Unexpected error:', e)

if __name__ == '__main__':
    main()